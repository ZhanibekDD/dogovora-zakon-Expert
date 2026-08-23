"""Exports every client/contract on record to an .xlsx spreadsheet.

One row per contract (a client with several contracts appears once per contract, so the sheet
reflects "who we did what for", not just a deduplicated client list). Columns: contract number,
date, client full name, IIN, phone, address, service type, amount, currency, payment type,
payment status, contract status.

Usage: python scripts/export_clients_xlsx.py [output_path.xlsx]
Defaults to storage/backups/clients_export_<timestamp>.xlsx if no path is given.
"""
from __future__ import annotations

import asyncio
import datetime
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession  # noqa: E402
from sqlalchemy.orm import selectinload  # noqa: E402

from app.core.config import get_settings  # noqa: E402
from app.database.models.contract import Contract  # noqa: E402
from app.database.session import session_scope  # noqa: E402

PAYMENT_TYPE_LABELS = {
    "prepayment": "предоплата",
    "after_result": "после результата",
    "split": "50/50",
    "already_paid": "уже оплачено",
    "custom": "по договорённости",
}

HEADERS = [
    "№ договора",
    "Дата",
    "ФИО клиента",
    "ИИН",
    "Телефон",
    "Адрес",
    "Услуга",
    "Сумма",
    "Валюта",
    "Оплата",
    "Статус оплаты",
    "Статус договора",
]


async def fetch_rows(session: AsyncSession) -> list[tuple]:
    result = await session.execute(
        select(Contract).options(selectinload(Contract.client)).order_by(Contract.contract_number)
    )
    contracts = result.scalars().all()

    rows = []
    for contract in contracts:
        client = contract.client
        service_type = (contract.service_data or {}).get("service_type") or ""
        created_at = contract.created_at
        date_str = created_at.strftime("%d.%m.%Y") if created_at else ""
        rows.append(
            (
                contract.contract_number,
                date_str,
                client.full_name if client else "",
                client.iin if client else "",
                client.phone if client else "",
                client.address if client else "",
                service_type,
                float(contract.amount or 0),
                contract.currency,
                PAYMENT_TYPE_LABELS.get(contract.payment_type, contract.payment_type),
                contract.payment_status,
                contract.status,
            )
        )
    return rows


def write_workbook(rows: list[tuple], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Договоры"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append(row)

    widths = [12, 12, 32, 14, 16, 30, 40, 12, 8, 18, 16, 16]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


async def main() -> None:
    settings = get_settings()
    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1])
    else:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = settings.backups_dir / f"clients_export_{timestamp}.xlsx"

    async with session_scope() as session:
        rows = await fetch_rows(session)
    write_workbook(rows, output_path)
    print(f"Exported {len(rows)} contract(s) to {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
